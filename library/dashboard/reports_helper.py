import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from io import BytesIO
from django.http import HttpResponse
from datetime import date

def generate_excel_report(report_type, queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{report_type.title()} Report"
    
    # Enable grid lines
    ws.views.sheetView[0].showGridLines = True
    
    # Styles
    title_font = Font(name='Segoe UI', size=16, bold=True, color='0F172A')
    header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
    data_font = Font(name='Segoe UI', size=10)
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    # Title Block
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = f"Lumina Portal - {report_type.replace('_', ' ').title()} Report"
    title_cell.font = title_font
    title_cell.alignment = left_align
    ws.row_dimensions[1].height = 40
    
    # Empty Row
    ws.append([])
    
    # Headers and Query compilation
    if report_type == 'books':
        headers = ['Title', 'Author(s)', 'Genre(s)', 'ISBN', 'Publisher', 'Publication Date', 'Total Copies', 'Available Copies']
    elif report_type == 'loans':
        headers = ['Borrower Username', 'Borrower Name', 'Book Title', 'ISBN', 'Borrow Date', 'Due Date', 'Status', 'Days Left / Overdue']
    elif report_type == 'fines':
        headers = ['Borrower Username', 'Borrower Name', 'Book Title', 'Due Date', 'Days Overdue', 'Fine Amount (INR)', 'Status']
    else:
        headers = []
        
    ws.append(headers)
    ws.row_dimensions[3].height = 26
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    # Data Rows
    row_num = 4
    if report_type == 'books':
        for book in queryset:
            authors = ", ".join([a.name for a in book.authors.all()])
            genres = ", ".join([g.name for g in book.genres.all()])
            pub_date = book.publication_date.strftime('%Y-%m-%d') if book.publication_date else 'N/A'
            row_data = [
                book.title, authors, genres, book.isbn, 
                book.publisher or 'N/A', pub_date, book.total_copies, book.available_copies
            ]
            ws.append(row_data)
            ws.row_dimensions[row_num].height = 20
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = center_align if col_idx in [4, 6, 7, 8] else left_align
            row_num += 1
            
    elif report_type == 'loans':
        for loan in queryset:
            name = loan.borrower.get_full_name() or loan.borrower.username
            status_display = loan.get_status_display()
            days_info = f"{loan.days_left} days left" if loan.days_left >= 0 else f"{-loan.days_left} days overdue"
            if loan.return_date:
                days_info = "Returned"
            row_data = [
                loan.borrower.username, name, loan.book.title, loan.book.isbn,
                loan.borrow_date.strftime('%Y-%m-%d'), loan.due_date.strftime('%Y-%m-%d'),
                status_display, days_info
            ]
            ws.append(row_data)
            ws.row_dimensions[row_num].height = 20
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = center_align if col_idx in [1, 4, 5, 6, 7, 8] else left_align
            row_num += 1
            
    elif report_type == 'fines':
        for loan in queryset:
            name = loan.borrower.get_full_name() or loan.borrower.username
            overdue_days = max(0, (date.today() - loan.due_date).days) if not loan.return_date else max(0, (loan.return_date - loan.due_date).days)
            status = "Paid" if loan.fine_paid else "Unpaid"
            row_data = [
                loan.borrower.username, name, loan.book.title,
                loan.due_date.strftime('%Y-%m-%d'), overdue_days, float(loan.fine_amount),
                status
            ]
            ws.append(row_data)
            ws.row_dimensions[row_num].height = 20
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = center_align if col_idx in [1, 4, 5, 7] else left_align
            row_num += 1
            
    # Autofit columns
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            # Skip title cell in column A when evaluating length for fit
            if cell.row == 1:
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = f'attachment; filename="lumina_{report_type}_report.xlsx"'
    wb.save(response)
    return response

def generate_pdf_report(report_type, queryset):
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=letter,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36
    )
    
    story = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=20,
        textColor=colors.HexColor('#0F172A'),
        spaceAfter=5
    )
    
    meta_style = ParagraphStyle(
        'DocMeta',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        textColor=colors.HexColor('#64748B'),
        spaceAfter=20
    )
    
    table_header_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        textColor=colors.white,
        alignment=1 # Center
    )
    
    table_cell_style = ParagraphStyle(
        'TableCell',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        textColor=colors.HexColor('#1E293B')
    )
    
    table_cell_center = ParagraphStyle(
        'TableCellCenter',
        parent=table_cell_style,
        alignment=1 # Center
    )
    
    # Title & Metadata
    story.append(Paragraph("Lumina Digital Library Portal", title_style))
    story.append(Paragraph(f"Report: {report_type.replace('_', ' ').title()} | Generated on: {date.today().strftime('%B %d, %Y')}", meta_style))
    story.append(Spacer(1, 10))
    
    # Prepare table structure
    if report_type == 'books':
        headers = ['Title', 'Author(s)', 'ISBN', 'Publisher', 'Total', 'Avail']
        col_widths = [160, 140, 70, 100, 35, 35] # Sum is 540 (printable width is 612 - 72 = 540)
        data = [[Paragraph(h, table_header_style) for h in headers]]
        for book in queryset:
            authors = ", ".join([a.name for a in book.authors.all()])
            data.append([
                Paragraph(book.title, table_cell_style),
                Paragraph(authors, table_cell_style),
                Paragraph(book.isbn, table_cell_center),
                Paragraph(book.publisher or 'N/A', table_cell_style),
                Paragraph(str(book.total_copies), table_cell_center),
                Paragraph(str(book.available_copies), table_cell_center),
            ])
            
    elif report_type == 'loans':
        headers = ['Borrower', 'Book Title', 'ISBN', 'Borrow Date', 'Due Date', 'Status']
        col_widths = [80, 180, 80, 70, 70, 60]
        data = [[Paragraph(h, table_header_style) for h in headers]]
        for loan in queryset:
            name = loan.borrower.get_full_name() or loan.borrower.username
            borrower_text = f"{loan.borrower.username}<br/><font color='#64748B' size='7'>{name}</font>"
            status_text = f"<font color='red'><b>{loan.status}</b></font>" if loan.status == 'OVERDUE' else loan.status
            data.append([
                Paragraph(borrower_text, table_cell_style),
                Paragraph(loan.book.title, table_cell_style),
                Paragraph(loan.book.isbn, table_cell_center),
                Paragraph(loan.borrow_date.strftime('%Y-%m-%d'), table_cell_center),
                Paragraph(loan.due_date.strftime('%Y-%m-%d'), table_cell_center),
                Paragraph(status_text, table_cell_center),
            ])
            
    elif report_type == 'fines':
        headers = ['Borrower', 'Book Title', 'Due Date', 'Overdue Days', 'Fine', 'Status']
        col_widths = [100, 200, 70, 60, 50, 60]
        data = [[Paragraph(h, table_header_style) for h in headers]]
        for loan in queryset:
            name = loan.borrower.get_full_name() or loan.borrower.username
            borrower_text = f"{loan.borrower.username}<br/><font color='#64748B' size='7'>{name}</font>"
            overdue_days = max(0, (date.today() - loan.due_date).days) if not loan.return_date else max(0, (loan.return_date - loan.due_date).days)
            status = "<font color='green'>Paid</font>" if loan.fine_paid else "<font color='red'>Unpaid</font>"
            data.append([
                Paragraph(borrower_text, table_cell_style),
                Paragraph(loan.book.title, table_cell_style),
                Paragraph(loan.due_date.strftime('%Y-%m-%d'), table_cell_center),
                Paragraph(str(overdue_days), table_cell_center),
                Paragraph(f"INR {loan.fine_amount:.2f}", table_cell_center),
                Paragraph(status, table_cell_center),
            ])
    else:
        col_widths = []
        data = []
            
    if data:
        table = Table(data, colWidths=col_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E293B')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('BOTTOMPADDING', (0,0), (-1,0), 8),
            ('TOPPADDING', (0,0), (-1,0), 8),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor('#F8FAFC'), colors.white]),
            ('TOPPADDING', (0,1), (-1,-1), 6),
            ('BOTTOMPADDING', (0,1), (-1,-1), 6),
        ]))
        story.append(table)
        
    doc.build(story)
    
    pdf = buffer.getvalue()
    buffer.close()
    
    response = HttpResponse(content_type="application/pdf")
    response['Content-Disposition'] = f'attachment; filename="lumina_{report_type}_report.pdf"'
    response.write(pdf)
    return response
